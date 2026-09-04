import io
import pytest
from openpyxl import load_workbook
from app.services.export_service import (
    export_excel,
    export_pdf,
    export_projects_list_excel,
    export_projects_list_pdf,
    _get_project_tipologia
)
from app.models.project import Project, ProjectStatus
from sqlalchemy.ext.asyncio import AsyncSession


def test_get_project_tipologia():
    class DummyProj:
        def __init__(self, is_atex=False, is_alimentare=False):
            self.is_atex = is_atex
            self.is_alimentare = is_alimentare

    assert _get_project_tipologia(DummyProj(False, False)) == "Standard"
    assert _get_project_tipologia(DummyProj(True, False)) == "ATEX"
    assert _get_project_tipologia(DummyProj(False, True)) == "Alimentare"
    assert _get_project_tipologia(DummyProj(True, True)) == "ATEX + Alimentare"


@pytest.mark.asyncio
async def test_export_excel_basic(db_session: AsyncSession, test_project: Project):
    result = await export_excel(db_session, test_project.id)
    assert hasattr(result, "getvalue")
    assert len(result.getvalue()) > 0

    # Load workbook and check sheet 1 contains Tipologia
    wb = load_workbook(filename=io.BytesIO(result.getvalue()))
    assert "Commessa" in wb.sheetnames
    ws_info = wb["Commessa"]
    row2_val = ws_info.cell(row=2, column=1).value or ""
    assert "Tipologia:" in row2_val
    assert "Standard" in row2_val


@pytest.mark.asyncio
async def test_export_excel_atex_alimentare(db_session: AsyncSession, test_project: Project):
    test_project.is_atex = True
    test_project.is_alimentare = True
    await db_session.commit()

    result = await export_excel(db_session, test_project.id)
    wb = load_workbook(filename=io.BytesIO(result.getvalue()))
    ws_info = wb["Commessa"]
    row2_val = ws_info.cell(row=2, column=1).value or ""
    assert "Tipologia: ATEX + Alimentare" in row2_val


@pytest.mark.asyncio
async def test_export_pdf_with_tipologia(db_session: AsyncSession, test_project: Project):
    test_project.is_atex = True
    test_project.is_alimentare = False
    await db_session.commit()

    pdf_buffer = await export_pdf(db_session, test_project.id)
    content = pdf_buffer.getvalue()
    assert len(content) > 0
    assert content.startswith(b"%PDF")


@pytest.mark.asyncio
async def test_export_projects_list_excel_and_pdf(db_session: AsyncSession, test_project: Project):
    excel_buf = await export_projects_list_excel(db_session, [test_project.id])
    wb = load_workbook(filename=io.BytesIO(excel_buf.getvalue()))
    ws = wb.active
    headers = [ws.cell(row=1, column=col).value for col in range(1, 9)]
    assert "Tipologia" in headers

    pdf_buf = await export_projects_list_pdf(db_session, [test_project.id])
    assert pdf_buf.getvalue().startswith(b"%PDF")

