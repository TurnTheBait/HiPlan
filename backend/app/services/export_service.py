import io
import json
from typing import List, Tuple, Optional, Any, cast
from datetime import timedelta, date, datetime
from app.utils.working_days import is_italian_holiday
# pyrefly: ignore [missing-import]
from sqlalchemy.ext.asyncio import AsyncSession
# pyrefly: ignore [missing-import]
from sqlalchemy import select
from app.models.task import Task
from app.models.project import Project
from app.models.ticket import Ticket, TicketReply
from app.models.user import User
# pyrefly: ignore [missing-import]
from reportlab.lib import colors
# pyrefly: ignore [missing-import]
from reportlab.lib.pagesizes import A4, landscape
# pyrefly: ignore [missing-import]
from reportlab.lib.units import mm
# pyrefly: ignore [missing-import]
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable, PageBreak
# pyrefly: ignore [missing-import]
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
# pyrefly: ignore [missing-import]
from reportlab.graphics.shapes import Drawing, Rect, Line, String
# pyrefly: ignore [missing-import]
from openpyxl import Workbook
# pyrefly: ignore [missing-import]
from openpyxl.worksheet.worksheet import Worksheet
# pyrefly: ignore [missing-import]
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# pyrefly: ignore [missing-import]
from openpyxl.utils import get_column_letter


def _extract_task_info(task: Any) -> Tuple[str, float, float, float]:
    """Estrae info base da un task: workers_str, planned, t_eff, diff"""
    workers_list = []
    if task.workers:
        try:
            parsed_w = json.loads(task.workers)
            if isinstance(parsed_w, list):
                workers_list = parsed_w
        except Exception:
            pass
    workers_str = ", ".join(workers_list) if workers_list else "-"

    t_eff = 0.0
    if task.actual_hours:
        try:
            parsed_h = json.loads(task.actual_hours)
            if isinstance(parsed_h, dict):
                for day_map in parsed_h.values():
                    if isinstance(day_map, dict):
                        for h in day_map.values():
                            try:
                                t_eff += float(h)
                            except (ValueError, TypeError):
                                pass
        except Exception:
            pass

    planned = float(task.planned_hours or 0.0)
    diff = t_eff - planned
    return workers_str, planned, t_eff, diff


def _get_workers_from_task(task: Any) -> List[str]:
    """Restituisce la lista di workers di un task"""
    try:
        parsed = json.loads(task.workers) if task.workers else []
        return parsed if isinstance(parsed, list) else []
    except:
        return []


def _get_worker_hours_map(task: Any) -> dict:
    """Restituisce il mappa worker_hours di un task"""
    try:
        return json.loads(task.worker_hours) if task.worker_hours else {}
    except:
        return {}


def _get_actual_hours_map(task: Any) -> dict:
    """Restituisce il mappa actual_hours di un task"""
    try:
        return json.loads(task.actual_hours) if task.actual_hours else {}
    except:
        return {}


def _get_task_color(task: Any) -> str:
    """Restituisce un colore esadecimale per il task"""
    if task.color:
        return str(task.color)
    colors_list = ["#3B82F6", "#10B981", "#F59E0B", "#EF4444", "#8B5CF6", "#EC4899", "#06B6D4", "#84CC16"]
    idx = hash(str(task.id or task.text)) % len(colors_list)
    return colors_list[idx]


def _compute_work_dates(start_date: Any, end_date: Any) -> List[Any]:
    """Restituisce lista di date lavorative (lun-ven) tra start e end"""
    dates = []
    if not start_date or not end_date:
        return dates
    cur = start_date
    while cur <= end_date:
        if cur.weekday() < 5:  # lun=0, ven=4
            dates.append(cur)
        cur += timedelta(days=1)
    return dates


async def _get_project_data(db: AsyncSession, project_id: str) -> Tuple[Any, List[Any]]:
    """Helper per fetchare progetto e tasks"""
    # pyrefly: ignore [missing-import]
    from sqlalchemy.orm import selectinload
    project = await db.execute(select(Project).options(selectinload(Project.responsible)).where(Project.id == project_id))
    proj = project.scalar_one()
    tasks = await db.execute(
        select(Task).where(Task.project_id == project_id).order_by(Task.sort_order)
    )
    task_list = list(tasks.scalars().all())
    return proj, task_list


# ===================== STILI COMUNI =====================
HEADER_FILL = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
SUMMARY_FONT = Font(bold=True, size=11, color="1E3A8A")
SUMMARY_FILL = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB")
)
SECTION_FILL = PatternFill(start_color="F0F4FF", end_color="F0F4FF", fill_type="solid")


# ==================== EXCEL EXPORT ====================

async def export_excel(db: AsyncSession, project_id: str, sections: Optional[List[str]] = None) -> io.BytesIO:
    if sections is None:
        sections = ["tasks", "hours", "gantt"]
    proj, task_list = await _get_project_data(db, project_id)
    wb = Workbook()

    # ========== SHEET 1: INFO COMMESSA (sempre presente) ==========
    ws_info = cast(Worksheet, wb.active)
    ws_info.title = "Commessa"

    ws_info.append([f"COMMESSA: {proj.code or ''} - {proj.name}"])
    ws_info.cell(row=1, column=1).font = Font(size=14, bold=True, color="1E3A8A")

    resp_str = proj.responsible.full_name or proj.responsible.username if proj.responsible else "-"
    addetti_list = []
    if proj.assigned_workers:
        try:
            parsed_aw = json.loads(proj.assigned_workers)
            if isinstance(parsed_aw, list):
                addetti_list = parsed_aw
        except:
            pass
    addetti_str = ", ".join(addetti_list) if addetti_list else "-"
    ws_info.append([f"Cliente: {proj.client or '-'} | Stato: {proj.status.value if proj.status else '-'}"])
    ws_info.cell(row=2, column=1).font = Font(size=11, italic=True)
    ws_info.append([f"Periodo: {proj.start_date or '-'} al {proj.end_date or '-'} | Responsabile: {resp_str}"])
    ws_info.cell(row=3, column=1).font = Font(size=11, italic=True)
    ws_info.append([f"Addetti Commessa: {addetti_str}"])
    ws_info.cell(row=4, column=1).font = Font(size=11, italic=True)
    ws_info.column_dimensions["A"].width = 120

    # ========== SHEET 2: GANTT DIAGRAM ==========
    if "gantt" in sections:
        ws_gantt = cast(Worksheet, wb.create_sheet("Diagramma Gantt"))
        has_tasks_section = "tasks" in sections
        has_hours_section = "hours" in sections

        # Find min/max dates: prioritize project boundaries
        min_date = proj.start_date
        max_date = proj.end_date

        if not min_date or not max_date:
            all_dates = []
            for t in task_list:
                if t.start_date:
                    all_dates.append(t.start_date)
                if t.end_date:
                    all_dates.append(t.end_date)
            
            if not min_date:
                min_date = min(all_dates) if all_dates else (proj.created_at.date() if hasattr(proj.created_at, 'date') else proj.created_at)
            if not max_date:
                max_date = max(all_dates) if all_dates else min_date
        
        # Generate list of days from min_date to max_date
        delta = max_date - min_date
        num_days = max(1, delta.days + 1)
        all_days = [min_date + timedelta(days=i) for i in range(num_days)]

        ws_gantt.append(["DIAGRAMMA GANTT - " + (proj.code or proj.name)])
        ws_gantt.cell(row=1, column=1).font = Font(size=14, bold=True, color="1E3A8A")
        ws_gantt.append([])

        # Row 3: Months header
        row = 3
        ws_gantt.cell(row=row, column=1, value="Commessa").font = Font(bold=True, size=10)
        ws_gantt.cell(row=row, column=1).fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
        ws_gantt.cell(row=row, column=1).border = THIN_BORDER

        col = 2
        from itertools import groupby
        for (year, month), g in groupby(all_days, key=lambda d: (d.year, d.month)):
            days_in_this_group = len(list(g))
            month_label = date(year, month, 1).strftime("%b %Y")
            ws_gantt.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + days_in_this_group - 1)
            cell = ws_gantt.cell(row=row, column=col, value=month_label)
            cell.font = Font(bold=True, size=9, color="2563EB")
            cell.fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
            col += days_in_this_group

        # Row 4 & 5: Day of week and day number headers
        row_dow = 4
        row_days = 5
        ws_gantt.cell(row=row_dow, column=1, value="").border = THIN_BORDER
        ws_gantt.cell(row=row_days, column=1, value="").border = THIN_BORDER
        
        ita_weekdays = ["Lun", "Mar", "Mer", "Gio", "Ven", "Sab", "Dom"]
        holiday_cols = set()
        col_days = 2
        for cur_date in all_days:
            is_festivo = is_italian_holiday(cur_date)
            
            if is_festivo:
                holiday_cols.add(col_days)
                
            bg_color = "FEE2E2" if is_festivo else "F9FAFB"
            text_color = "DC2626" if is_festivo else "6B7280"
            
            cell_dow = ws_gantt.cell(row=row_dow, column=col_days, value=ita_weekdays[cur_date.weekday()])
            cell_dow.font = Font(size=8, color=text_color, bold=is_festivo)
            cell_dow.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            cell_dow.border = THIN_BORDER
            cell_dow.alignment = Alignment(horizontal="center")
            
            cell_day = ws_gantt.cell(row=row_days, column=col_days, value=f"{cur_date.day:02d}")
            cell_day.font = Font(size=8, color=text_color, bold=is_festivo)
            cell_day.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            cell_day.border = THIN_BORDER
            cell_day.alignment = Alignment(horizontal="center")
            col_days += 1

        # Determine total columns used
        total_cols = col_days - 1

        # Project bar
        row = 6
        ws_gantt.cell(row=row, column=1, value=f"{proj.code or ''} {proj.name}").font = Font(size=9, bold=True)
        ws_gantt.cell(row=row, column=1).border = THIN_BORDER
        
        for c in range(2, total_cols + 1):
            bg = "FEE2E2" if c in holiday_cols else "FFFFFF"
            cell_bg = ws_gantt.cell(row=row, column=c)
            cell_bg.border = THIN_BORDER
            cell_bg.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        
        if proj.start_date and proj.end_date:
            p_start = proj.start_date
            p_end = proj.end_date
            start_offset = (p_start - min_date).days if p_start >= min_date else 0
            end_offset = (p_end - min_date).days if p_end >= min_date else 0

            bar_width = max(1, end_offset - start_offset + 1)
            start_col = 2 + start_offset
            end_col = min(start_col + bar_width - 1, total_cols)
            
            if end_col >= start_col:
                ws_gantt.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
                bar_cell = ws_gantt.cell(row=row, column=start_col, value=f"{proj.code or ''} {proj.name}")
                bar_cell.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
                bar_cell.font = Font(color="FFFFFF", bold=True, size=8)
                bar_cell.border = THIN_BORDER
                bar_cell.alignment = Alignment(horizontal="center")

        # Task bars
        row = 7
        for task in task_list:
            ws_gantt.cell(row=row, column=1, value=task.text).font = Font(size=9)
            ws_gantt.cell(row=row, column=1).border = THIN_BORDER
            ws_gantt.cell(row=row, column=1).fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

            for c in range(2, total_cols + 1):
                bg = "FEE2E2" if c in holiday_cols else "FFFFFF"
                cell_bg = ws_gantt.cell(row=row, column=c)
                cell_bg.border = THIN_BORDER
                cell_bg.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")

            if task.start_date and task.end_date:
                t_start = task.start_date
                t_end = task.end_date
                start_offset = (t_start - min_date).days if t_start >= min_date else 0
                end_offset = (t_end - min_date).days if t_end >= min_date else 0

                bar_width = max(1, end_offset - start_offset + 1)
                start_col = 2 + start_offset
                end_col = min(start_col + bar_width - 1, total_cols)
                
                ws_gantt.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
                bar_cell = ws_gantt.cell(row=row, column=start_col, value=task.text)
                t_color = _get_task_color(task)
                bar_cell.fill = PatternFill(start_color=t_color[1:], end_color=t_color[1:], fill_type="solid")
                bar_cell.font = Font(color="FFFFFF", bold=True, size=8)
                bar_cell.border = THIN_BORDER
                bar_cell.alignment = Alignment(horizontal="center")

            row += 1

        ws_gantt.column_dimensions["A"].width = 40
        for c in range(2, total_cols + 2):
            ws_gantt.column_dimensions[get_column_letter(c)].width = 4

    # ========== SHEET 3: TABELLA FASI ==========
    if "tasks" in sections:
        ws_tasks = cast(Worksheet, wb.create_sheet("Fasi"))
        ws_tasks.append(["TABELLA FASI"])
        ws_tasks.cell(row=1, column=1).font = Font(size=14, bold=True, color="1E3A8A")
        ws_tasks.append([])

        has_hours_t = "hours" in sections

        headers = ["#", "Fase / Attività", "Inizio", "Fine", "Durata (gg)"]
        if has_hours_t:
            headers += ["Ore Budget (h)", "Ore Consuntivate (h)", "Saldo Ore (h)"]
        headers += ["Reparto", "Addetti Assegnati", "Progresso %", "Priorità"]
        ws_tasks.append(headers)
        header_row = ws_tasks.max_row

        for c in range(1, len(headers) + 1):
            cell = ws_tasks.cell(row=header_row, column=c)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER

        tot_days = 0
        tot_planned = 0.0
        tot_actual = 0.0

        for i, task in enumerate(task_list, 1):
            workers_str, planned, t_eff, diff = _extract_task_info(task)
            tot_days += int(task.duration or 0)
            tot_planned += planned
            tot_actual += t_eff

            row_data = [i, task.text,
                        str(task.start_date) if task.start_date else "",
                        str(task.end_date) if task.end_date else "",
                        task.duration]
            if has_hours_t:
                row_data += [round(planned, 1), round(t_eff, 1), round(diff, 1)]
            row_data += [task.department if task.department else "-",
                        workers_str,
                        f"{task.progress * 100:.0f}%",
                        task.priority.value if task.priority else "-"]
            ws_tasks.append(row_data)
            current_row = ws_tasks.max_row
            for ci in range(1, len(row_data) + 1):
                c = ws_tasks.cell(row=current_row, column=ci)
                c.border = THIN_BORDER
                c.alignment = Alignment(horizontal="center", vertical="center")

        # Total row
        ws_tasks.append([])
        tot_diff = tot_actual - tot_planned
        tot_row = ["", "TOTALE COMMESSA:", "", "", tot_days]
        if has_hours_t:
            tot_row += [round(tot_planned, 1), round(tot_actual, 1), round(tot_diff, 1)]
        tot_row += ["", "", "", ""]
        ws_tasks.append(tot_row)
        tot_row_idx = ws_tasks.max_row
        for ci in range(1, len(tot_row) + 1):
            c = ws_tasks.cell(row=tot_row_idx, column=ci)
            c.fill = SUMMARY_FILL
            c.font = SUMMARY_FONT
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal="center", vertical="center")

        ws_tasks.column_dimensions["A"].width = 6
        ws_tasks.column_dimensions["B"].width = 45
        ws_tasks.column_dimensions["C"].width = 13
        ws_tasks.column_dimensions["D"].width = 13
        ws_tasks.column_dimensions["E"].width = 13
        
        col_offset = 5
        if has_hours_t:
            ws_tasks.column_dimensions["F"].width = 16
            ws_tasks.column_dimensions["G"].width = 20
            ws_tasks.column_dimensions["H"].width = 15
            col_offset = 8
            
        ws_tasks.column_dimensions[get_column_letter(col_offset + 1)].width = 20  # Reparto
        ws_tasks.column_dimensions[get_column_letter(col_offset + 2)].width = 30  # Addetti
        ws_tasks.column_dimensions[get_column_letter(col_offset + 3)].width = 15  # Progresso
        ws_tasks.column_dimensions[get_column_letter(col_offset + 4)].width = 15  # Priorità

    # ========== SHEET 3 o 4: CONSUNTIVO ORE ==========
    if "hours" in sections:
        ws_hours = cast(Worksheet, wb.create_sheet("Consuntivo Ore"))

        ws_hours.append(["CONSUNTIVAZIONE ORE PER ADDETTO"])
        ws_hours.cell(row=1, column=1).font = Font(size=14, bold=True, color="1E3A8A")
        ws_hours.append([])

        current_row = 3

        for task in task_list:
            workers = _get_workers_from_task(task)
            if not workers:
                workers = ["Addetto Generico"]
            actual_hours = _get_actual_hours_map(task)
            
            base_dates = _compute_work_dates(task.start_date, task.end_date)
            date_set = set(base_dates)
            for w, dates_dict in actual_hours.items():
                if isinstance(dates_dict, dict):
                    for date_str in dates_dict.keys():
                        try:
                            extra_d = datetime.strptime(date_str, "%Y-%m-%d").date()
                            date_set.add(extra_d)
                        except ValueError:
                            pass
            work_dates = sorted(list(date_set))
            
            worker_hours_map = _get_worker_hours_map(task)
            planned_total = float(task.planned_hours or 8.0)

            if not work_dates:
                continue

            # Task header
            ws_hours.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5 + len(work_dates))
            cell = ws_hours.cell(row=current_row, column=1,
                                value=f"📋 {task.text} — {planned_total}h previste")
            cell.font = Font(size=12, bold=True, color="1E3A8A")
            cell.fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
            cell.border = THIN_BORDER
            current_row += 1

            # Header row: Addetto | Ore Totali | Giorno1 | Giorno2 | ...
            header_row_data = ["Addetto", "Ore Totali Assegnate"]
            for d in work_dates:
                header_row_data.append(d.strftime("%d/%m"))
            ws_hours.append(header_row_data)
            hr = ws_hours.max_row
            for ci in range(1, len(header_row_data) + 1):
                c = ws_hours.cell(row=hr, column=ci)
                c.fill = HEADER_FILL
                c.font = HEADER_FONT
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                c.border = THIN_BORDER

            # Data rows per worker
            for w in workers:
                assigned_h = float(worker_hours_map.get(w, planned_total / len(workers))) if worker_hours_map else planned_total / len(workers)
                hours_per_day = assigned_h / len(work_dates) if work_dates else assigned_h

                row_data = [w, round(assigned_h, 1)]
                worker_total_actual = 0
                for d in work_dates:
                    date_str = d.strftime("%Y-%m-%d")
                    val = 0
                    if w in actual_hours and date_str in actual_hours[w]:
                        try:
                            val = float(actual_hours[w][date_str])
                        except:
                            val = 0
                    worker_total_actual += val
                    row_data.append(round(val, 1) if val else "-")

                ws_hours.append(row_data)
                dr = ws_hours.max_row
                for ci in range(1, len(row_data) + 1):
                    c = ws_hours.cell(row=dr, column=ci)
                    c.border = THIN_BORDER
                    c.alignment = Alignment(horizontal="center", vertical="center")
                    if ci == 1:
                        c.font = Font(bold=True)
                current_row = ws_hours.max_row + 1

            # Separator
            ws_hours.append([])
            current_row = ws_hours.max_row + 1

        # Column widths
        ws_hours.column_dimensions["A"].width = 22
        ws_hours.column_dimensions["B"].width = 22
        for ci in range(3, 50):
            col_letter = get_column_letter(ci)
            ws_hours.column_dimensions[col_letter].width = 12

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ==================== PDF EXPORT ====================

async def export_pdf(db: AsyncSession, project_id: str, sections: Optional[List[str]] = None) -> io.BytesIO:
    if sections is None:
        sections = ["tasks", "hours", "gantt"]
    proj, task_list = await _get_project_data(db, project_id)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        topMargin=12 * mm, bottomMargin=12 * mm,
        leftMargin=12 * mm, rightMargin=12 * mm
    )
    styles = getSampleStyleSheet()
    elements = []

    # Stili
    title_style = ParagraphStyle('TitleCustom', parent=styles['Title'], fontSize=16, textColor=colors.HexColor("#1E3A8A"), spaceAfter=6)
    section_title = ParagraphStyle('SectionTitle', parent=styles['Heading2'], fontSize=14, textColor=colors.HexColor("#2563EB"), spaceBefore=6, spaceAfter=8)
    cell_left = ParagraphStyle('CellLeft', parent=styles['Normal'], fontName='Helvetica', fontSize=8, leading=10, textColor=colors.HexColor('#1e293b'))
    cell_center = ParagraphStyle('CellCenter', parent=cell_left, alignment=1)
    cell_bold_center = ParagraphStyle('CellBoldCenter', parent=cell_center, fontName='Helvetica-Bold')
    cell_bold_left = ParagraphStyle('CellBoldLeft', parent=cell_left, fontName='Helvetica-Bold')
    header_style = ParagraphStyle('HeaderCell', parent=cell_center, fontName='Helvetica-Bold', fontSize=8, leading=10, textColor=colors.white)
    gantt_bar_style = ParagraphStyle('GanttBar', parent=cell_center, fontName='Helvetica-Bold', fontSize=6, textColor=colors.white, leading=8)
    gantt_label_style = ParagraphStyle('GanttLabel', parent=cell_left, fontName='Helvetica', fontSize=7, leading=9)

    # ===== INTESTAZIONE (solo sulla prima pagina) =====
    elements.append(Paragraph(f"<b>{proj.code or ''} {proj.name}</b> — Report Commessa", title_style))

    resp_str = proj.responsible.full_name or proj.responsible.username if proj.responsible else "-"
    addetti_list = []
    if proj.assigned_workers:
        try:
            parsed_aw = json.loads(proj.assigned_workers)
            if isinstance(parsed_aw, list):
                addetti_list = parsed_aw
        except: pass
    addetti_str = ", ".join(addetti_list) if addetti_list else "-"

    info_str = (f"<b>Cliente:</b> {proj.client or '-'} | <b>Stato:</b> {proj.status.value if proj.status else '-'} | "
                f"<b>Periodo:</b> {proj.start_date or '-'} → {proj.end_date or '-'}<br/>"
                f"<b>Responsabile:</b> {resp_str} | <b>Addetti:</b> {addetti_str}")
    if proj.description:
        info_str += f"<br/><b>Descrizione:</b> {proj.description}"
    elements.append(Paragraph(info_str, styles["Normal"]))
    elements.append(Spacer(1, 4 * mm))

    # ===== SEZIONE 1: GANTT =====
    # ===== SEZIONE 1: GANTT =====
    if "gantt" in sections:
        elements.append(PageBreak())
        elements.append(Paragraph("📊 Diagramma Gantt", section_title))

        # Find date range across project and tasks
        all_dates = []
        if proj.start_date: all_dates.append(proj.start_date)
        if proj.end_date: all_dates.append(proj.end_date)
        for t in task_list:
            if t.start_date: all_dates.append(t.start_date)
            if t.end_date: all_dates.append(t.end_date)
        if not all_dates:
            all_dates = [proj.start_date or proj.created_at, proj.end_date or proj.start_date]
        min_date = min(all_dates)
        max_date = max(all_dates)

        # Assicura che min_date parta dall'inizio del primo mese e max_date alla fine dell'ultimo
        min_date = min_date.replace(day=1)
        if max_date.month == 12:
            max_date = max_date.replace(year=max_date.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            max_date = max_date.replace(month=max_date.month + 1, day=1) - timedelta(days=1)

        total_days = (max_date - min_date).days + 1
        if total_days <= 0:
            total_days = 1

        chart_width = 600  # pt for timeline drawing area
        label_width = 172  # pt for task label column (totale tabella 772 pt, perfetto per A4 orizzontale)

        # Build month info list
        current = min_date.replace(day=1)
        months_info = []
        mesi_it = {
            1: "Gennaio", 2: "Febbraio", 3: "Marzo", 4: "Aprile",
            5: "Maggio", 6: "Giugno", 7: "Luglio", 8: "Agosto",
            9: "Settembre", 10: "Ottobre", 11: "Novembre", 12: "Dicembre"
        }
        while current <= max_date:
            m_label = f"{mesi_it[current.month]} {current.year}"
            if current.month == 12:
                next_m = current.replace(year=current.year + 1, month=1)
            else:
                next_m = current.replace(month=current.month + 1)
            month_end = min(next_m - timedelta(days=1), max_date)
            months_info.append((current, month_end, m_label))
            current = next_m

        # Header row drawing
        header_drawing = Drawing(chart_width, 24)
        header_drawing.add(Rect(0, 0, chart_width, 24, fillColor=colors.HexColor("#1E40AF"), strokeColor=colors.HexColor("#1E40AF"), strokeWidth=0))
        for m_start, m_end, m_label in months_info:
            x1 = (m_start - min_date).days * chart_width / total_days
            x2 = ((m_end - min_date).days + 1) * chart_width / total_days
            w = max(0.5, x2 - x1)
            header_drawing.add(Rect(x1, 0, w, 24, fillColor=colors.HexColor("#1E40AF"), strokeColor=colors.HexColor("#93C5FD"), strokeWidth=0.5))
            if w < 45:
                display_label = f"{mesi_it[m_start.month][:3]} '{str(m_start.year)[-2:]}"
            else:
                display_label = m_label
            header_drawing.add(String(x1 + w / 2, 7, display_label, fontName="Helvetica-Bold", fontSize=8, fillColor=colors.white, textAnchor="middle"))

        gantt_data = [
            [Paragraph("<b>Fase / Attività</b>", header_style), header_drawing]
        ]

        # Helper to draw vertical gridlines on row drawings
        def add_row_gridlines(dw):
            for m_s, m_e, _ in months_info:
                gx = (m_s - min_date).days * chart_width / total_days
                dw.add(Line(gx, 0, gx, 22, strokeColor=colors.HexColor("#CBD5E1"), strokeWidth=0.5))
            dw.add(Line(chart_width, 0, chart_width, 22, strokeColor=colors.HexColor("#CBD5E1"), strokeWidth=0.5))

        # Project bar row
        if proj.start_date and proj.end_date:
            proj_drawing = Drawing(chart_width, 22)
            add_row_gridlines(proj_drawing)
            p_left = max(0, min(total_days, (proj.start_date - min_date).days)) * chart_width / total_days
            p_right = max(0, min(total_days, (proj.end_date - min_date).days + 1)) * chart_width / total_days
            bar_w = max(4, p_right - p_left)
            proj_drawing.add(Rect(p_left, 4, bar_w, 14, rx=3, ry=3, fillColor=colors.HexColor("#2563EB"), strokeColor=colors.HexColor("#1D4ED8"), strokeWidth=0.5))
            if bar_w >= 65:
                proj_drawing.add(String(p_left + bar_w / 2, 7, f"{proj.code or ''} {proj.name}", fontName="Helvetica-Bold", fontSize=7, fillColor=colors.white, textAnchor="middle"))
            gantt_data.append([
                Paragraph(f"<b>{proj.code or ''} {proj.name}</b>", gantt_label_style),
                proj_drawing
            ])

        # Task rows
        for task in task_list:
            if not task.start_date or not task.end_date:
                continue
            row_drawing = Drawing(chart_width, 22)
            add_row_gridlines(row_drawing)
            t_color = _get_task_color(task)
            t_left = max(0, min(total_days, (task.start_date - min_date).days)) * chart_width / total_days
            t_right = max(0, min(total_days, (task.end_date - min_date).days + 1)) * chart_width / total_days
            bar_w = max(4, t_right - t_left)
            row_drawing.add(Rect(t_left, 4, bar_w, 14, rx=3, ry=3, fillColor=colors.HexColor(t_color), strokeColor=colors.HexColor(t_color), strokeWidth=0.5))

            duration_days = max(1, (task.end_date - task.start_date).days + 1)
            dur_str = f"{task.duration}g" if task.duration else f"{duration_days}g"
            if bar_w >= 35:
                row_drawing.add(String(t_left + bar_w / 2, 7, dur_str, fontName="Helvetica-Bold", fontSize=6.5, fillColor=colors.white, textAnchor="middle"))
            elif t_left + bar_w + 25 <= chart_width:
                row_drawing.add(String(t_left + bar_w + 4, 7, dur_str, fontName="Helvetica", fontSize=6.5, fillColor=colors.HexColor("#475569"), textAnchor="start"))

            gantt_data.append([
                Paragraph(task.text, gantt_label_style),
                row_drawing
            ])

        gantt_table = Table(gantt_data, colWidths=[label_width, chart_width], repeatRows=1)
        style_cmds = [
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CBD5E1")),
            ("BACKGROUND", (0, 0), (0, 0), colors.HexColor("#1E40AF")),
            ("BACKGROUND", (1, 0), (1, 0), colors.HexColor("#1E40AF")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ("LEFTPADDING", (0, 0), (0, -1), 6),
            ("RIGHTPADDING", (0, 0), (0, -1), 6),
            ("LEFTPADDING", (1, 0), (1, -1), 0),
            ("RIGHTPADDING", (1, 0), (1, -1), 0),
        ]
        if proj.start_date and proj.end_date:
            style_cmds.append(("BACKGROUND", (0, 1), (0, 1), colors.HexColor("#EFF6FF")))

        start_ri = 2 if (proj.start_date and proj.end_date) else 1
        for idx in range(start_ri, len(gantt_data)):
            if idx % 2 == 1:
                style_cmds.append(("BACKGROUND", (0, idx), (0, idx), colors.HexColor("#F8FAFC")))

        gantt_table.setStyle(TableStyle(style_cmds))
        elements.append(gantt_table)
        elements.append(Spacer(1, 4 * mm))

    # ===== SEZIONE 2: TABELLA FASI =====
    if "tasks" in sections:
        elements.append(PageBreak())
        elements.append(Paragraph("📋 Dettaglio Fasi", section_title))

        has_hours_t = "hours" in sections

        tot_days_pdf = 0
        tot_planned_pdf = 0.0
        tot_actual_pdf = 0.0
        for task in task_list:
            ws, planned, t_eff, diff = _extract_task_info(task)
            tot_days_pdf += int(task.duration or 0)
            tot_planned_pdf += planned
            tot_actual_pdf += t_eff

        # KPI summary
        tot_diff_pdf = tot_actual_pdf - tot_planned_pdf
        kpi_data = [
            [Paragraph("<b>Fasi Totali</b>", cell_center),
             Paragraph("<b>Durata Stimata</b>", cell_center),
             Paragraph("<b>Budget Ore</b>", cell_center),
             Paragraph("<b>Ore Consuntivate</b>", cell_center),
             Paragraph("<b>Saldo Ore</b>", cell_center)],
            [Paragraph(f"<b>{len(task_list)}</b>", cell_center),
             Paragraph(f"<b>{tot_days_pdf} giorni</b>", cell_center),
             Paragraph(f"<b>{tot_planned_pdf:.1f} h</b>", cell_center),
             Paragraph(f"<b>{tot_actual_pdf:.1f} h</b>", cell_center),
             Paragraph(f"<b>{tot_diff_pdf:+.1f} h</b>", cell_center)]
        ]
        kpi_table = Table(kpi_data, colWidths=[120, 120, 140, 140, 140])
        kpi_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EFF6FF")),
            ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#DBEAFE")),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#93C5FD")),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(kpi_table)
        elements.append(Spacer(1, 4 * mm))

        # Main tasks table
        headers = [Paragraph("#", header_style), Paragraph("Fase / Attività", header_style),
                   Paragraph("Inizio", header_style), Paragraph("Fine", header_style),
                   Paragraph("Giorni", header_style)]
        if has_hours_t:
            headers += [Paragraph("Ore<br/>Budget", header_style), Paragraph("Ore<br/>Effett.", header_style),
                        Paragraph("Saldo<br/>Ore", header_style)]
        headers += [Paragraph("Reparto", header_style), Paragraph("Addetti", header_style),
                    Paragraph("Prog.", header_style), Paragraph("Priorità", header_style)]

        data = [headers]
        for i, task in enumerate(task_list, 1):
            ws, planned, t_eff, diff = _extract_task_info(task)
            diff_str = f"{diff:+.1f}h" if abs(diff) > 0.05 else "0.0h"
            row = [Paragraph(str(i), cell_center), Paragraph(task.text, cell_left),
                   Paragraph(str(task.start_date) if task.start_date else "-", cell_center),
                   Paragraph(str(task.end_date) if task.end_date else "-", cell_center),
                   Paragraph(f"{task.duration}g", cell_center)]
            if has_hours_t:
                row += [Paragraph(f"{planned:.1f}h", cell_center), Paragraph(f"{t_eff:.1f}h", cell_center),
                        Paragraph(diff_str, cell_center)]
            row += [Paragraph(task.department if task.department else "-", cell_center),
                    Paragraph(ws, cell_left),
                    Paragraph(f"{task.progress * 100:.0f}%", cell_center),
                    Paragraph(task.priority.value if task.priority else "-", cell_center)]
            data.append(row)

        # Total row
        tot_diff_str = f"{tot_diff_pdf:+.1f}h" if abs(tot_diff_pdf) > 0.05 else "0.0h"
        tot_row = [Paragraph("", cell_center), Paragraph("<b>TOTALE</b>", cell_bold_left),
                   Paragraph("", cell_center), Paragraph("", cell_center),
                   Paragraph(f"<b>{tot_days_pdf}g</b>", cell_bold_center)]
        if has_hours_t:
            tot_row += [Paragraph(f"<b>{tot_planned_pdf:.1f}h</b>", cell_bold_center),
                        Paragraph(f"<b>{tot_actual_pdf:.1f}h</b>", cell_bold_center),
                        Paragraph(f"<b>{tot_diff_str}</b>", cell_bold_center)]
        tot_row += [Paragraph("", cell_center), Paragraph("", cell_left),
                    Paragraph("", cell_center), Paragraph("", cell_center)]
        data.append(tot_row)

        col_widths = [18, 120, 50, 50, 35]
        if has_hours_t:
            col_widths += [45, 45, 45]
        col_widths += [60, 130, 38, 50]
        
        table = Table(data, colWidths=col_widths)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#F8FAFC")]),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#EFF6FF")),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 6 * mm))

    # ===== SEZIONE 3: CONSUNTIVO ORE =====
    if "hours" in sections:
        elements.append(PageBreak())
        elements.append(Paragraph("⏱ Consuntivazione Ore per Addetto", section_title))

        for task in task_list:
            workers = _get_workers_from_task(task)
            if not workers:
                workers = ["Addetto Generico"]
            actual_hours = _get_actual_hours_map(task)
            
            base_dates = _compute_work_dates(task.start_date, task.end_date)
            date_set = set(base_dates)
            for w, dates_dict in actual_hours.items():
                if isinstance(dates_dict, dict):
                    for date_str in dates_dict.keys():
                        try:
                            extra_d = datetime.strptime(date_str, "%Y-%m-%d").date()
                            date_set.add(extra_d)
                        except ValueError:
                            pass
            work_dates = sorted(list(date_set))
            
            worker_hours_map = _get_worker_hours_map(task)
            planned_total = float(task.planned_hours or 8.0)

            if not work_dates:
                continue

            # Task title
            elements.append(Paragraph(
                f"<b>{task.text}</b> — {planned_total}h previste ({len(workers)} addetti, {len(work_dates)} giorni lavorativi)",
                ParagraphStyle('TaskLabel', parent=styles['Normal'], fontSize=9, textColor=colors.HexColor("#1E3A8A"),
                               spaceBefore=8, spaceAfter=4, backColor=colors.HexColor("#EFF6FF"))
            ))

            # Header
            headers = [Paragraph("Addetto", header_style), Paragraph("Ore Assegnate", header_style)]
            for d in work_dates:
                headers.append(Paragraph(d.strftime("%d/%m"), header_style))

            data = [headers]
            for w in workers:
                assigned_h = float(worker_hours_map.get(w, planned_total / len(workers))) if worker_hours_map else planned_total / len(workers)
                row = [Paragraph(w, cell_center), Paragraph(f"{assigned_h:.1f}h", cell_center)]
                for d in work_dates:
                    date_str = d.strftime("%Y-%m-%d")
                    val = 0
                    if w in actual_hours and date_str in actual_hours[w]:
                        try:
                            val = float(actual_hours[w][date_str])
                        except: val = 0
                    row.append(Paragraph(f"{val:.1f}" if val else "-", cell_center))
                data.append(row)

            date_col_widths = [32] * len(work_dates)
            col_widths = [70, 60] + date_col_widths
            # Limit width to not exceed page
            max_width = 750
            if sum(col_widths) > max_width:
                col_widths = [70, 60] + [max(20, min(32, (max_width - 130) // len(work_dates)))] * len(work_dates)

            table = Table(data, colWidths=col_widths)
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#059669")),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D1D5DB")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0FDF4")]),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ]))
            elements.append(table)
            elements.append(Spacer(1, 4 * mm))

    doc.build(elements)
    buffer.seek(0)
    return buffer

async def export_projects_list_excel(db: AsyncSession, project_ids: List[str]) -> io.BytesIO:
    from app.models.user import User
    buffer = io.BytesIO()
    wb = Workbook()
    ws = cast(Worksheet, wb.active)
    ws.title = "Elenco Commesse"

    headers = ["Codice", "Nome Commessa", "Cliente", "Responsabile", "Stato", "Avanzamento", "Addetti"]
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    if not project_ids:
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    res = await db.execute(select(Project).where(Project.id.in_(project_ids)))
    projects = res.scalars().all()
    
    user_res = await db.execute(select(User))
    users = user_res.scalars().all()
    user_map = {u.id: u.full_name or u.username for u in users}

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for idx, p in enumerate(projects, 2):
        res_tasks = await db.execute(select(Task).where(Task.project_id == p.id))
        tasks = res_tasks.scalars().all()
        
        total_tasks = len(tasks)
        completed_tasks = sum(1 for t in tasks if t.completed)
        progress = f"{int((completed_tasks / total_tasks) * 100)}%" if total_tasks > 0 else "0%"
        
        workers = set()
        for t in tasks:
            workers.update(_get_workers_from_task(t))
        workers_str = ", ".join(sorted(workers)) if workers else "-"
        
        status_it = {"planning": "Pianificazione", "active": "In Corso", "completed": "Completata", "archived": "Archiviata"}.get(p.status, p.status)
        resp_name = user_map.get(p.responsible_id, "-") if p.responsible_id else "-"

        row_data = [
            p.code or "-",
            p.name,
            p.client or "-",
            resp_name,
            status_it,
            progress,
            workers_str
        ]
        
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=idx, column=col_idx, value=val)
            cell.border = thin_border
            if col_idx in [1, 4, 5, 6]:
                cell.alignment = Alignment(horizontal="center")

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["G"].width = 40

    wb.save(buffer)
    buffer.seek(0)
    return buffer

async def export_projects_list_pdf(db: AsyncSession, project_ids: List[str]) -> io.BytesIO:
    from app.models.user import User
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=10*mm, leftMargin=10*mm, topMargin=15*mm, bottomMargin=15*mm)
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle("TitleStyle", parent=styles["Heading1"], textColor=colors.HexColor("#1F2937"), fontSize=16, spaceAfter=10)
    elements.append(Paragraph("Elenco Commesse & Progetti", title_style))
    elements.append(Spacer(1, 5*mm))

    if not project_ids:
        doc.build(elements)
        buffer.seek(0)
        return buffer

    res = await db.execute(select(Project).where(Project.id.in_(project_ids)))
    projects = res.scalars().all()
    
    user_res = await db.execute(select(User))
    users = user_res.scalars().all()
    user_map = {u.id: u.full_name or u.username for u in users}

    header_style = ParagraphStyle("HeaderStyle", parent=styles["Normal"], textColor=colors.whitesmoke, fontName="Helvetica-Bold")
    
    data = [
        [
            Paragraph("Codice", header_style),
            Paragraph("Nome Commessa", header_style),
            Paragraph("Cliente", header_style),
            Paragraph("Resp.", header_style),
            Paragraph("Stato", header_style),
            Paragraph("Avanz.", header_style),
            Paragraph("Addetti", header_style)
        ]
    ]
    
    cell_style = ParagraphStyle("CellStyle", parent=styles["Normal"], fontSize=10, leading=12)
    
    for p in projects:
        res_tasks = await db.execute(select(Task).where(Task.project_id == p.id))
        tasks = res_tasks.scalars().all()
        
        total_tasks = len(tasks)
        completed_tasks = sum(1 for t in tasks if t.completed)
        progress = f"{int((completed_tasks / total_tasks) * 100)}%" if total_tasks > 0 else "0%"
        
        workers = set()
        for t in tasks:
            workers.update(_get_workers_from_task(t))
        workers_str = ", ".join(sorted(workers)) if workers else "-"
        
        status_it = {"planning": "Pianificazione", "active": "In Corso", "completed": "Completata", "archived": "Archiviata"}.get(p.status, p.status)
        resp_name = user_map.get(p.responsible_id, "-") if p.responsible_id else "-"
        
        data.append([
            Paragraph(p.code or "-", cell_style),
            Paragraph(p.name, cell_style),
            Paragraph(p.client or "-", cell_style),
            Paragraph(resp_name, cell_style),
            Paragraph(status_it, cell_style),
            Paragraph(progress, cell_style),
            Paragraph(workers_str, cell_style)
        ])

    table = Table(data, colWidths=[65, 160, 110, 85, 75, 45, 190])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("ALIGN", (5, 0), (5, -1), "CENTER"), # Avanzamento
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("TOPPADDING", (0, 0), (-1, 0), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D1D5DB")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
    ]))
    
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer


async def export_ticket_excel(db: AsyncSession, ticket_id: str) -> io.BytesIO:
    # pyrefly: ignore [missing-import]
    from sqlalchemy.orm import selectinload
    res = await db.execute(select(Ticket).options(selectinload(Ticket.project)).where(Ticket.id == ticket_id))
    ticket = res.scalar_one_or_none()
    if not ticket:
        raise ValueError("Ticket non trovato")

    res_replies = await db.execute(select(TicketReply).where(TicketReply.ticket_id == ticket_id).order_by(TicketReply.created_at))
    replies = res_replies.scalars().all()

    # Get users for author/responsible resolution
    res_users = await db.execute(select(User))
    users = res_users.scalars().all()
    user_map = {u.id: (u.full_name or u.username) for u in users}

    wb = Workbook()
    
    # --- Foglio 1: Dettagli Ticket ---
    ws_ticket = wb.active
    ws_ticket.title = "Dettagli Ticket"
    
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    
    # Headers
    headers = ["Campo", "Valore"]
    for col_num, header in enumerate(headers, 1):
        cell = ws_ticket.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill

    t_atts = []
    if ticket.attachments:
        try:
            parsed = json.loads(ticket.attachments)
            t_atts = [a.get("name", "Allegato") for a in parsed]
        except Exception:
            pass
    t_atts_str = ", ".join(t_atts) if t_atts else "-"

    details = [
        ("Codice / Progetto", ticket.custom_project_code or (ticket.project.name if ticket.project else "-")),
        ("Titolo", ticket.title),
        ("Stato", ticket.status),
        ("Priorità", ticket.priority),
        ("Creato Da", user_map.get(ticket.author_id, "-")),
        ("Responsabile", user_map.get(ticket.responsible_id, "-")),
        ("Descrizione", ticket.description or "-"),
        ("Allegati", t_atts_str),
    ]

    for row_num, (campo, valore) in enumerate(details, 2):
        ws_ticket.cell(row=row_num, column=1, value=campo)
        ws_ticket.cell(row=row_num, column=2, value=valore)

    ws_ticket.column_dimensions['A'].width = 20
    ws_ticket.column_dimensions['B'].width = 60

    # --- Foglio 2: Risposte e Timeline ---
    ws_replies = wb.create_sheet("Risposte e Timeline")
    replies_headers = ["Data", "Autore", "Tipo Azione", "Contenuto"]
    
    for col_num, header in enumerate(replies_headers, 1):
        cell = ws_replies.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for row_num, reply in enumerate(replies, 2):
        r_atts = []
        if reply.attachments:
            try:
                parsed = json.loads(reply.attachments)
                r_atts = [a.get("name", "Allegato") for a in parsed]
            except Exception:
                pass
        content_text = reply.content or ""
        if r_atts:
            content_text += f"\n[Allegati: {', '.join(r_atts)}]"
            
        ws_replies.cell(row=row_num, column=1, value=reply.created_at.strftime("%Y-%m-%d %H:%M"))
        ws_replies.cell(row=row_num, column=2, value=user_map.get(reply.author_id, "-"))
        ws_replies.cell(row=row_num, column=3, value=reply.action_type or "Nota Interna")
        ws_replies.cell(row=row_num, column=4, value=content_text)

    ws_replies.column_dimensions['A'].width = 20
    ws_replies.column_dimensions['B'].width = 25
    ws_replies.column_dimensions['C'].width = 25
    ws_replies.column_dimensions['D'].width = 80

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


async def export_ticket_pdf(db: AsyncSession, ticket_id: str) -> io.BytesIO:
    # pyrefly: ignore [missing-import]
    from sqlalchemy.orm import selectinload
    res = await db.execute(select(Ticket).options(selectinload(Ticket.project)).where(Ticket.id == ticket_id))
    ticket = res.scalar_one_or_none()
    if not ticket:
        raise ValueError("Ticket non trovato")

    res_replies = await db.execute(select(TicketReply).where(TicketReply.ticket_id == ticket_id).order_by(TicketReply.created_at))
    replies = res_replies.scalars().all()

    # Get users for author/responsible resolution
    res_users = await db.execute(select(User))
    users = res_users.scalars().all()
    user_map = {u.id: (u.full_name or u.username) for u in users}

    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=A4,
        rightMargin=15*mm,
        leftMargin=15*mm,
        topMargin=15*mm,
        bottomMargin=15*mm
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("TitleStyle", parent=styles["Heading1"], fontSize=18, textColor=colors.HexColor("#111827"), spaceAfter=12)
    normal_style = styles["Normal"]
    bold_style = ParagraphStyle("BoldStyle", parent=styles["Normal"], fontName="Helvetica-Bold")

    elements = []

    # Title
    elements.append(Paragraph(f"Ticket: {ticket.title}", title_style))
    elements.append(Spacer(1, 10))

    t_atts = []
    if ticket.attachments:
        try:
            parsed = json.loads(ticket.attachments)
            t_atts = [a.get("name", "Allegato") for a in parsed]
        except Exception:
            pass
    t_atts_str = ", ".join(t_atts) if t_atts else "-"

    # Details Table
    details_data = [
        [Paragraph("<b>Progetto:</b>", normal_style), Paragraph(ticket.custom_project_code or (ticket.project.name if ticket.project else "-"), normal_style)],
        [Paragraph("<b>Stato:</b>", normal_style), Paragraph(ticket.status, normal_style)],
        [Paragraph("<b>Priorità:</b>", normal_style), Paragraph(ticket.priority, normal_style)],
        [Paragraph("<b>Autore:</b>", normal_style), Paragraph(user_map.get(ticket.author_id, "-"), normal_style)],
        [Paragraph("<b>Responsabile:</b>", normal_style), Paragraph(user_map.get(ticket.responsible_id, "-"), normal_style)],
        [Paragraph("<b>Allegati:</b>", normal_style), Paragraph(t_atts_str, normal_style)],
    ]
    
    details_table = Table(details_data, colWidths=[100, 350], hAlign="LEFT")
    details_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(details_table)
    elements.append(Spacer(1, 15))

    elements.append(Paragraph("<b>Descrizione:</b>", bold_style))
    elements.append(Paragraph(ticket.description or "Nessuna descrizione fornita.", normal_style))
    elements.append(Spacer(1, 20))
    elements.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#E5E7EB"), spaceBefore=10, spaceAfter=20))

    # Replies Table
    elements.append(Paragraph("Cronologia Risposte", ParagraphStyle("H2", parent=styles["Heading2"], spaceAfter=10)))

    replies_data = [
        [Paragraph("Data", bold_style), Paragraph("Autore", bold_style), Paragraph("Azione", bold_style), Paragraph("Contenuto", bold_style)]
    ]

    for reply in replies:
        r_atts = []
        if reply.attachments:
            try:
                parsed = json.loads(reply.attachments)
                r_atts = [a.get("name", "Allegato") for a in parsed]
            except Exception:
                pass
                
        content_text = (reply.content or "").replace('\n', '<br/>')
        if r_atts:
            content_text += f"<br/><i>[Allegati: {', '.join(r_atts)}]</i>"

        replies_data.append([
            Paragraph(reply.created_at.strftime("%Y-%m-%d %H:%M"), normal_style),
            Paragraph(user_map.get(reply.author_id, "-"), normal_style),
            Paragraph(reply.action_type or "Nota Interna", normal_style),
            Paragraph(content_text, normal_style),
        ])

    replies_table = Table(replies_data, colWidths=[80, 100, 100, 240])
    replies_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("TOPPADDING", (0, 0), (-1, 0), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D1D5DB")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
    ]))
    elements.append(replies_table)

    doc.build(elements)
    output.seek(0)
    return output

def _is_vacation(date_str, user_dict):
    for v in user_dict.get("vacations", []):
        if v.get("start_date") and v.get("end_date"):
            if v["start_date"] <= date_str <= v["end_date"]:
                return True
    return False

async def export_workload_excel(heatmap_data: dict, mode: str = "both") -> io.BytesIO:
    # pyrefly: ignore [missing-import]
    import openpyxl
    import datetime
    # pyrefly: ignore [missing-import]
    from openpyxl.styles import Font, Alignment, PatternFill
    from app.utils.working_days import is_italian_holiday
    
    wb = openpyxl.Workbook()
    
    sheets = []
    if mode in ["planned", "both"]:
        ws_planned = wb.active if len(sheets) == 0 else wb.create_sheet()
        ws_planned.title = "Ore Pianificate"
        sheets.append((ws_planned, "planned"))
    
    if mode in ["actual", "both"]:
        ws_actual = wb.active if len(sheets) == 0 else wb.create_sheet()
        ws_actual.title = "Ore Consuntivate"
        sheets.append((ws_actual, "actual"))
    
    today = datetime.date.today()
    start_date = today - datetime.timedelta(days=365)
    end_date = today + datetime.timedelta(days=365)
    sorted_dates = [(start_date + datetime.timedelta(days=i)).strftime("%Y-%m-%d") for i in range((end_date - start_date).days + 1)]
    
    headers = ["Addetto", "Reparto"] + sorted_dates
    for ws, _ in sheets:
        ws.append(headers)
    
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    holiday_fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
    holiday_col_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for ws, _ in sheets:
        for col_idx, h_text in enumerate(headers):
            cell = ws.cell(row=1, column=col_idx + 1)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if col_idx >= 2:
                d = datetime.datetime.strptime(h_text, "%Y-%m-%d").date()
                if is_italian_holiday(d):
                    cell.fill = holiday_fill
                else:
                    cell.fill = header_fill
            else:
                cell.fill = header_fill
        
    for user_id, u in heatmap_data.items():
        workload = u.get("workload", {})
        actual_workload = u.get("actual_workload", {})
        
        for ws, sheet_type in sheets:
            row = [u.get("full_name") or u.get("username"), u.get("department") or "-"]
            
            for d in sorted_dates:
                prev_data = workload.get(d, {})
                eff_data = actual_workload.get(d, {})
                prev_val = prev_data.get("hours", 0) if isinstance(prev_data, dict) else 0
                eff_val = eff_data.get("hours", 0) if isinstance(eff_data, dict) else 0
                
                p_val = float(prev_val)
                e_val = float(eff_val)
                
                if _is_vacation(d, u):
                    if sheet_type == "planned":
                        row.append("Ferie" if p_val == 0 else f"Ferie ({p_val})")
                    else:
                        row.append("Ferie" if e_val == 0 else f"Ferie ({e_val})")
                else:
                    if sheet_type == "planned":
                        row.append(p_val)
                    else:
                        row.append(e_val)
                        
            ws.append(row)
        
    for ws, _ in sheets:
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(3, len(headers) + 1):
                d_str = headers[col_idx - 1]
                d = datetime.datetime.strptime(d_str, "%Y-%m-%d").date()
                if is_italian_holiday(d):
                    ws.cell(row=row_idx, column=col_idx).fill = holiday_col_fill
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

async def export_workload_pdf(heatmap_data: dict, mode: str = "both") -> io.BytesIO:
    # pyrefly: ignore [missing-import]
    from reportlab.lib import colors
    # pyrefly: ignore [missing-import]
    from reportlab.lib.pagesizes import A4, landscape
    # pyrefly: ignore [missing-import]
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    # pyrefly: ignore [missing-import]
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    import datetime
    from collections import defaultdict
    
    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=15,
        leftMargin=15,
        topMargin=15,
        bottomMargin=15
    )
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title", parent=styles["Heading1"], fontSize=16, spaceAfter=10)
    elements = []
    
    title_text = "Saturazione Carichi di Lavoro (Pianificate vs Consuntivate)"
    if mode == "planned":
        title_text = "Saturazione Carichi di Lavoro (Ore Pianificate)"
    elif mode == "actual":
        title_text = "Saturazione Carichi di Lavoro (Ore Consuntivate)"
        
    elements.append(Paragraph(title_text, title_style))
    elements.append(Spacer(1, 10))
    
    today = datetime.date.today()
    start_date = today - datetime.timedelta(days=365)
    end_date = today + datetime.timedelta(days=365)
    
    week_map = {}
    week_to_month = {}
    for i in range((end_date - start_date).days + 1):
        d = start_date + datetime.timedelta(days=i)
        iso_year, iso_week, _ = d.isocalendar()
        key = (iso_year, iso_week)
        if key not in week_map:
            week_map[key] = f"W{iso_week:02d}"
            # Use Thursday to determine the month of the ISO week
            thursday = d + datetime.timedelta(days=(3 - d.weekday()))
            week_to_month[key] = (thursday.year, thursday.month)
            
    sorted_week_keys = sorted(week_map.keys())
    
    user_weeks = {}
    for uid, u in heatmap_data.items():
        user_weeks[uid] = defaultdict(lambda: {"prev": 0.0, "eff": 0.0})
        
        for d_str, data in u.get("workload", {}).items():
            if d_str != '__extra__':
                dt = datetime.datetime.strptime(d_str, "%Y-%m-%d").date()
                if start_date <= dt <= end_date:
                    iso_year, iso_week, _ = dt.isocalendar()
                    key = (iso_year, iso_week)
                    val = data.get("hours", 0) if isinstance(data, dict) else 0
                    user_weeks[uid][key]["prev"] += float(val)
                    
        for d_str, data in u.get("actual_workload", {}).items():
            if d_str != '__extra__':
                dt = datetime.datetime.strptime(d_str, "%Y-%m-%d").date()
                if start_date <= dt <= end_date:
                    iso_year, iso_week, _ = dt.isocalendar()
                    key = (iso_year, iso_week)
                    val = data.get("hours", 0) if isinstance(data, dict) else 0
                    user_weeks[uid][key]["eff"] += float(val)
                    
    month_to_weeks = defaultdict(list)
    for key in sorted_week_keys:
        month_to_weeks[week_to_month[key]].append(key)
        
    sorted_months = sorted(month_to_weeks.keys())
    month_names = ["Gennaio", "Febbraio", "Marzo", "Aprile", "Maggio", "Giugno", "Luglio", "Agosto", "Settembre", "Ottobre", "Novembre", "Dicembre"]
    
    for m_year, m_month in sorted_months:
        week_keys = month_to_weeks[(m_year, m_month)]
        m_name = month_names[m_month - 1]
        
        elements.append(Paragraph(f"{m_name} {m_year}", ParagraphStyle("Sub", parent=styles["Heading2"], spaceAfter=10)))
        
        headers = ["Addetto", "Reparto"] + [week_map[k] for k in week_keys]
        sub_data = [headers]
        
        for uid, u in heatmap_data.items():
            row = [u.get("full_name") or u.get("username"), u.get("department") or "-"]
            for k in week_keys:
                prev = user_weeks[uid][k]["prev"]
                eff = user_weeks[uid][k]["eff"]
                
                if mode == "planned":
                    row.append(f"{prev:.1f} h")
                elif mode == "actual":
                    row.append(f"{eff:.1f} h")
                else:
                    row.append(f"{eff:.1f} / {prev:.1f} h")
            sub_data.append(row)
            
        t = Table(sub_data)
        style_cmds = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D1D5DB")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ]
        t.setStyle(TableStyle(style_cmds))
        elements.append(t)
        elements.append(Spacer(1, 20))
        
    doc.build(elements)
    output.seek(0)
    return output
